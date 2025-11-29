# -*- coding: utf-8 -*-
import sys
import json
import pandas as pd
from openpyxl import load_workbook

# Configurar codificación UTF-8 para stdout
if sys.stdout.encoding != 'utf-8':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def write_row(file_path, sheet_name, row_data):
    """Add a new row to a specific sheet"""
    try:
        # Load workbook
        wb = load_workbook(file_path)
        
        # Get or create sheet
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]
        
        # Find the next empty row
        next_row = ws.max_row + 1
        
        # Write data
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=next_row, column=col_idx, value=value)
        
        # Save workbook
        wb.save(file_path)
        wb.close()
        
        return {
            'success': True,
            'message': 'Row added successfully',
            'row': next_row
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

def update_row(file_path, sheet_name, row_index, row_data):
    """Update a specific row in a sheet"""
    try:
        wb = load_workbook(file_path)
        
        if sheet_name not in wb.sheetnames:
            return {'success': False, 'error': f'Sheet {sheet_name} not found'}
        
        ws = wb[sheet_name]
        
        # Update row
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_index, column=col_idx, value=value)
        
        wb.save(file_path)
        wb.close()
        
        return {
            'success': True,
            'message': 'Row updated successfully'
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

def delete_row(file_path, sheet_name, row_index):
    """Delete a specific row from a sheet"""
    try:
        wb = load_workbook(file_path)
        
        if sheet_name not in wb.sheetnames:
            return {'success': False, 'error': f'Sheet {sheet_name} not found'}
        
        ws = wb[sheet_name]
        
        # Delete row
        ws.delete_rows(row_index)
        
        wb.save(file_path)
        wb.close()
        
        return {
            'success': True,
            'message': 'Row deleted successfully'
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

if __name__ == '__main__':
    if len(sys.argv) < 4:
        print(json.dumps({'success': False, 'error': 'Missing arguments'}))
        sys.exit(1)
    
    command = sys.argv[1]
    file_path = sys.argv[2]
    sheet_name = sys.argv[3]
    
    if command == 'create':
        if len(sys.argv) < 5:
            result = {'success': False, 'error': 'Missing row data'}
        else:
            row_data = json.loads(sys.argv[4])
            result = write_row(file_path, sheet_name, row_data)
    elif command == 'update':
        if len(sys.argv) < 6:
            result = {'success': False, 'error': 'Missing row index or data'}
        else:
            row_index = int(sys.argv[4])
            row_data = json.loads(sys.argv[5])
            result = update_row(file_path, sheet_name, row_index, row_data)
    elif command == 'delete':
        if len(sys.argv) < 5:
            result = {'success': False, 'error': 'Missing row index'}
        else:
            row_index = int(sys.argv[4])
            result = delete_row(file_path, sheet_name, row_index)
    else:
        result = {'success': False, 'error': 'Unknown command'}
    
    print(json.dumps(result, ensure_ascii=False))

